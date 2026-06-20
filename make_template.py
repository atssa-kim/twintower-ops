"""
트윈타워 운영앱 — 데이터 입력 Excel 템플릿 생성기
출력: twintower_template.xlsx  (시트 11장)
"""
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── 공통 스타일 ──────────────────────────────────────────────
def side(style="thin"):
    return Side(style=style, color="AAAAAA")

BORDER = Border(left=side(), right=side(), top=side(), bottom=side())
BORDER_BOLD = Border(left=side("medium"), right=side("medium"),
                     top=side("medium"), bottom=side("medium"))

def hdr(cell, text, bg="1F4E79", fg="FFFFFF", size=11, bold=True, wrap=False):
    cell.value = text
    cell.font = Font(name="맑은 고딕", bold=bold, color=fg, size=size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                               wrap_text=wrap)
    cell.border = BORDER_BOLD

def cell(ws, row, col, text, bg=None, bold=False, color="000000",
         align="left", wrap=False, size=10):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name="맑은 고딕", bold=bold, color=color, size=size)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center",
                            wrap_text=wrap)
    c.border = BORDER
    return c

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def freeze(ws, cell_addr):
    ws.freeze_panes = cell_addr

def note(ws, row, text, col=1, span=8):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name="맑은 고딕", italic=True, color="555555", size=9)
    c.alignment = Alignment(horizontal="left", vertical="center")


# ════════════════════════════════════════════════════════════
# 시트 1 — 명부 (roster)
# ════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "①명부(roster)"
ws1.row_dimensions[1].height = 30
ws1.row_dimensions[2].height = 20
ws1.row_dimensions[3].height = 22

# 제목
ws1.merge_cells("A1:I1")
c = ws1["A1"]
c.value = "트윈타워 전체 근무자 명부 (roster)"
c.font = Font(name="맑은 고딕", bold=True, size=14, color="FFFFFF")
c.fill = PatternFill("solid", fgColor="1F4E79")
c.alignment = Alignment(horizontal="center", vertical="center")

# 안내
ws1.merge_cells("A2:I2")
note(ws1, 2, "※ 이 시트가 모든 재난·근무표의 기준입니다. 사번은 고유값, 변경 금지.")

# 헤더
headers1 = ["사번", "이름", "파트", "직책", "근무유형\n(일근/교대)", "교대조\n(교대자만)",
            "연락처", "FCM토큰\n(앱 로그인 후 자동입력)", "비고"]
for i, h in enumerate(headers1, 1):
    hdr(ws1.cell(row=3, column=i), h, wrap=True)

# 예시 데이터
examples = [
    ("B-0001", "센터장명", "-",      "시설관리센터장",    "일근", "",       "010-0000-0000", "", ""),
    ("B-1001", "소방파트장", "소방파트", "소방안전관리자",  "교대", "소방1조", "010-0000-0000", "", ""),
    ("B-1002", "소방대원",  "소방파트", "소방대원",       "교대", "소방2조", "010-0000-0000", "", ""),
    ("B-2001", "전기파트장", "전기파트", "전기파트장",     "일근", "",       "010-0000-0000", "", ""),
    ("B-2002", "전기대원",  "전기파트", "전기대원",       "교대", "BMS1조",  "010-0000-0000", "", ""),
    ("B-3001", "기계파트장", "기계파트", "기계파트장",     "일근", "",       "010-0000-0000", "", ""),
    ("B-4001", "건축파트장", "건축파트", "건축파트장",     "일근", "",       "010-0000-0000", "", ""),
    ("B-5001", "운영파트장", "운영파트", "운영파트장",     "일근", "",       "010-0000-0000", "", ""),
    ("B-6001", "보안실장",  "보안파트", "보안실장",       "교대", "보안1조", "010-0000-0000", "", ""),
    ("B-7001", "품질파트장", "품질파트", "품질파트장",     "일근", "",       "010-0000-0000", "", ""),
    ("B-8001", "주차담당",  "주차파트", "주차담당",       "교대", "주차1조", "010-0000-0000", "", ""),
    ("B-9001", "미화담당",  "미화파트", "미화담당",       "일근", "",       "010-0000-0000", "", ""),
]
row_colors = ["FFFFFF", "EEF4FF"]
for r, ex in enumerate(examples, 4):
    bg = row_colors[r % 2]
    for c_idx, val in enumerate(ex, 1):
        cell(ws1, r, c_idx, val, bg=bg, align="center" if c_idx in [1,5,6,7] else "left")

# 범례
note_row = 4 + len(examples) + 1
ws1.merge_cells(f"A{note_row}:I{note_row}")
note(ws1, note_row,
     "▶ 파트 목록: 소방파트/전기파트/기계파트/건축파트/운영파트/품질파트/보안파트/주차파트/미화파트")
ws1.merge_cells(f"A{note_row+1}:I{note_row+1}")
note(ws1, note_row+1,
     "▶ 교대조 예시: 소방1~4조 / BMS1~4조 / 보안1~4조 / 주차1~2조 등 (실제 조 편성에 맞게 기입)")
ws1.merge_cells(f"A{note_row+2}:I{note_row+2}")
note(ws1, note_row+2,
     "▶ FCM토큰: 앱 첫 로그인 시 자동 등록 — 직접 입력 불필요")

set_col_widths(ws1, [12, 12, 12, 18, 14, 12, 16, 28, 14])
freeze(ws1, "A4")


# ════════════════════════════════════════════════════════════
# 시트 2 — 교대 근무표 (shifts)
# ════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("②근무표(shifts)")
ws2.row_dimensions[1].height = 30
ws2.row_dimensions[2].height = 20
ws2.row_dimensions[3].height = 22
ws2.row_dimensions[4].height = 20

ws2.merge_cells("A1:AK1")
c = ws2["A1"]
c.value = "교대 근무표 — 2026년 __월  (교대 근무자만 입력)"
c.font = Font(name="맑은 고딕", bold=True, size=14, color="FFFFFF")
c.fill = PatternFill("solid", fgColor="375623")
c.alignment = Alignment(horizontal="center", vertical="center")

ws2.merge_cells("A2:AK2")
note(ws2, 2, "※ 근무코드: 주=주간(08:00~18:00) / 야=야간(18:00~익일08:00) / 아=아침(별도정의) / 비=비번(알람제외)")

# 고정 헤더
fixed = ["사번", "이름", "파트", "교대조"]
for i, h in enumerate(fixed, 1):
    hdr(ws2.cell(row=3, column=i), h, bg="375623")

# 날짜 헤더 (1~31일)
days_of_week = ["월","화","수","목","금","토","일"]
# 2026-06-01은 월요일 (weekday=0)
import datetime
base = datetime.date(2026, 6, 1)
for d in range(1, 32):
    col = 4 + d
    dt = datetime.date(2026, 6, d) if d <= 30 else None
    label = str(d)
    bg = "375623"
    if dt:
        dow = dt.weekday()
        if dow == 5:   bg = "C00000"  # 토 빨강
        elif dow == 6: bg = "C00000"  # 일 빨강
    c3 = ws2.cell(row=3, column=col)
    hdr(c3, label, bg=bg, size=9)
    if dt:
        c4 = ws2.cell(row=4, column=col)
        dow_label = days_of_week[dt.weekday()]
        hdr(c4, dow_label, bg="C00000" if dt.weekday() >= 5 else "4A7C59",
            size=9)

ws2.merge_cells("A4:D4")
note(ws2, 4, "← 6월 날짜/요일 →", col=1)

# 예시 데이터
shift_examples = [
    ("B-1001", "김상백", "소방파트", "소방1조",
     "비주주야비주주야비주주야비주주야비주주야비주주야비주주야비주주야비"),
    ("B-1002", "박범수", "소방파트", "소방2조",
     "주야비주주야비주주야비주주야비주주야비주주야비주주야비주주야비주주"),
    ("B-1003", "안준혁", "소방파트", "소방3조",
     "야비주주야비주주야비주주야비주주야비주주야비주주야비주주야비주주야"),
    ("B-1004", "김병기", "소방파트", "소방4조",
     "주주야비주주야비주주야비주주야비주주야비주주야비주주야비주주야비주"),
    ("B-2001", "BMS조장1", "전기파트", "BMS1조",
     "비주주야비주주야비주주야비주주야비주주야비주주야비주주야비주주야비"),
    ("B-2002", "BMS조원1", "전기파트", "BMS2조",
     "주야비주주야비주주야비주주야비주주야비주주야비주주야비주주야비주주"),
]
CODE_COLORS = {"주": "DDEEFF", "야": "FFF0CC", "비": "F5F5F5", "아": "E8FFE8"}

for r, ex in enumerate(shift_examples, 5):
    sno, name, dept, grp, codes = ex
    cell(ws2, r, 1, sno,  bg="FFFFFF", align="center")
    cell(ws2, r, 2, name, bg="FFFFFF")
    cell(ws2, r, 3, dept, bg="FFFFFF")
    cell(ws2, r, 4, grp,  bg="FFFFFF", align="center")
    for d, code in enumerate(codes[:30], 1):
        col = 4 + d
        bg = CODE_COLORS.get(code, "FFFFFF")
        cell(ws2, r, col, code, bg=bg, align="center", bold=(code != "비"),
             color="C00000" if code == "야" else "1F4E79" if code == "주" else "555555")

# 빈 입력 행
for r in range(5 + len(shift_examples), 5 + len(shift_examples) + 10):
    for col in range(1, 36):
        cell(ws2, r, col, "", bg="FAFAFA")

note_r = 5 + len(shift_examples) + 11
ws2.merge_cells(f"A{note_r}:AK{note_r}")
note(ws2, note_r, "▶ 시트를 복사하여 월별로 관리 (예: ②근무표_2026-07, ②근무표_2026-08 ...)")

col_widths2 = [12, 10, 10, 10] + [4] * 31
set_col_widths(ws2, col_widths2)
freeze(ws2, "E3")


# ════════════════════════════════════════════════════════════
# 헬퍼: 재난별 조직도 시트 생성
# ════════════════════════════════════════════════════════════
DISASTER_COLOR = {
    "화재":     ("C00000", "FF0000"),
    "정전":     ("7B6100", "FFC000"),
    "누수·침수": ("1F4E79", "4472C4"),
    "풍수해":   ("1F4E79", "70AD47"),
    "폭설":     ("2E4057", "9DC3E6"),
    "지진":     ("5C3A00", "ED7D31"),
    "가스누출":  ("375623", "70AD47"),
    "승강기갇힘":("4A235A", "AA64FF"),
    "테러·침입": ("404040", "FF4444"),
}

GROUP_COLORS = {
    "지휘":     "FFF2CC",
    "총괄관리자":"FFF2CC",
    "연락반":   "DDEBF7",
    "연락":     "DDEBF7",
    "상황":     "DDEBF7",
    "상황실":   "DDEBF7",
    "대응반":   "FCE4D6",
    "유도반":   "E2EFDA",
    "대피반":   "E2EFDA",
    "복구반":   "F2F2F2",
    "지원반":   "F2F2F2",
}

def make_org_sheet(wb, disaster_name, sheet_num, rows):
    """
    rows: list of (반, 역할, 담당파트, 사번1, 이름1, 사번2, 이름2, ...)
    """
    title = f"③조직도_{sheet_num}_{disaster_name}"
    ws = wb.create_sheet(title)
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 22

    bg_dark, bg_light = DISASTER_COLOR.get(disaster_name, ("1F4E79", "4472C4"))

    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = f"재난 조직도 — {disaster_name}  (orgCharts/{disaster_name.replace('·','_')})"
    c.font = Font(name="맑은 고딕", bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=bg_dark)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:J2")
    note(ws, 2, "※ 사번은 명부(①시트)의 사번과 반드시 일치. 한 역할에 여러 명이면 행 추가.")

    hdr_cols = ["반\n(group)", "역할\n(role)", "담당파트", "사번①", "이름①",
                "사번②", "이름②", "사번③", "이름③", "비고"]
    for i, h in enumerate(hdr_cols, 1):
        hdr(ws.cell(row=3, column=i), h, bg=bg_dark, wrap=True)

    for r_idx, row_data in enumerate(rows, 4):
        group = row_data[0]
        bg = GROUP_COLORS.get(group, "FFFFFF")
        for c_idx, val in enumerate(row_data, 1):
            bold = (c_idx == 1)
            cell(ws, r_idx, c_idx, val, bg=bg, bold=bold,
                 align="center" if c_idx in [1,4,5,6,7,8,9] else "left")
        # 빈 셀 채우기
        for c_idx in range(len(row_data)+1, 11):
            cell(ws, r_idx, c_idx, "", bg=bg)

    # 빈 입력 행
    for r_idx in range(4+len(rows), 4+len(rows)+5):
        for c_idx in range(1, 11):
            cell(ws, r_idx, c_idx, "", bg="FAFAFA")

    set_col_widths(ws, [12, 28, 20, 12, 12, 12, 12, 12, 12, 14])
    freeze(ws, "A4")
    return ws


# 재난별 조직도 데이터 (반, 역할, 담당파트, 사번①, 이름①, ...)
ORG_DATA = {
    "화재": [
        ("지휘",  "총괄자 (센터장)",            "-",            "B-0001","센터장명","","","","",""),
        ("지휘",  "소방안전관리자 (소방파트장)",  "소방파트",      "B-1001","소방파트장","","","","",""),
        ("연락반","상황실 (통보연락)",            "소방파트/BMS",  "B-2001","상황실①","B-2002","상황실②","","",""),
        ("대응반","비상출동조",                   "전기파트/소방파트","B-3001","전기①","B-3002","소방①","B-3003","소방②",""),
        ("대응반","소화조",                       "기계파트",      "B-4001","기계①","B-4002","기계②","","",""),
        ("대응반","인명구조조",                   "보안파트",      "B-6001","보안①","B-6002","보안②","","",""),
        ("유도반","유도조",                       "운영/보안/주차","B-5001","운영①","B-5002","보안③","B-8001","주차①",""),
        ("유도반","경계조",                       "보안파트",      "B-6003","보안④","","","","",""),
        ("복구반","복구조",                       "미화/시설",     "B-9001","미화①","","","","",""),
    ],
    "정전": [
        ("지휘",  "총괄관리자 (센터장)",          "-",            "B-0001","센터장명","","","","",""),
        ("연락",  "상황실 (조장)",                "소방파트/BMS",  "B-2001","상황실①","","","","",""),
        ("대응반","통제자 (전기파트장)",           "전기파트",      "B-3001","전기파트장","","","","",""),
        ("대응반","대응조 (전기파트)",             "전기파트",      "B-3002","전기①","B-3003","전기②","","",""),
        ("지원반","지원조 (소방·기계)",            "소방파트/기계파트","B-1001","소방①","B-4001","기계①","","",""),
        ("지원반","유도·지원반 (운영·보안)",        "운영파트/보안파트","B-5001","운영①","B-6001","보안①","","",""),
    ],
    "누수·침수": [
        ("총괄관리자","총괄관리자 — 센터장",      "-",            "B-0001","센터장명","","","","",""),
        ("상황실","상황실 — 당직조장",            "소방파트/BMS",  "B-2001","조장①","","","","",""),
        ("대응반","응급조 (기계/소방)",            "기계파트/소방파트","B-4001","기계①","B-1001","소방①","","",""),
        ("대응반","복구조(전기)",                  "전기파트",      "B-3001","전기①","B-3002","전기②","","",""),
        ("대응반","복구조(건축/미화/협력)",         "건축파트/미화파트","B-7001","건축①","B-9001","미화①","","",""),
        ("대피반","유도조 (보안/협력)",            "보안파트",      "B-6001","보안①","B-6002","보안②","","",""),
        ("대피반","보안조 (운영파트장)",           "운영파트",      "B-5001","운영파트장","","","","",""),
        ("지원반","의료조",                        "보안파트",      "B-6003","보안③","","","","",""),
        ("지원반","관리조 (운영파트)",             "운영파트",      "B-5002","운영①","","","","",""),
    ],
    "풍수해": [
        ("지휘",  "총괄관리자 (센터장)",           "-",           "B-0001","센터장명","","","","",""),
        ("상황",  "상황실 (조장)",                 "소방파트/BMS", "B-2001","상황실①","","","","",""),
        ("대응반","통제자 (건축파트장)",            "건축파트",     "B-7001","건축파트장","","","","",""),
        ("대응반","대응조 (건축·기계·전기·운영)",    "각 파트",     "B-7002","건축①","B-4001","기계①","B-3001","전기①",""),
        ("지원반","지원조 (미화·보안·주차)",         "미화/보안/주차","B-9001","미화①","B-6001","보안①","B-8001","주차①",""),
        ("지원반","구조조",                         "보안파트",    "B-6002","보안②","","","","",""),
    ],
    "폭설": [
        ("지휘",  "총괄관리자 (센터장)",            "-",          "B-0001","센터장명","","","","",""),
        ("대응반","대응1조 (운영·소방·건축)",        "각 파트",    "B-5001","운영①","B-1001","소방①","B-7001","건축①",""),
        ("대응반","대응2조 (전기·기계·품질)",        "각 파트",    "B-3001","전기①","B-4001","기계①","B-2001","품질①",""),
        ("지원반","지원1조 (미화협력)",              "미화파트",   "B-9001","미화①","B-9002","미화②","","",""),
        ("지원반","지원2조 (보안·주차협력)",          "보안/주차",  "B-6001","보안①","B-8001","주차①","","",""),
    ],
    "지진": [
        ("지휘",  "총괄관리자 (센터장)",            "-",          "B-0001","센터장명","","","","",""),
        ("대응반","대응1조 (전기·건축)",             "전기/건축",  "B-3001","전기①","B-7001","건축①","","",""),
        ("대응반","대응2조 (기계·미화·주차)",         "기계/미화/주차","B-4001","기계①","B-9001","미화①","B-8001","주차①",""),
        ("대피반","대피반 (보안파트)",               "보안파트",   "B-6001","보안①","B-6002","보안②","","",""),
        ("지원반","지원반 (운영파트)",               "운영파트",   "B-5001","운영①","B-5002","운영②","","",""),
    ],
    "가스누출": [
        ("지휘",  "총괄자 (센터장)",                "-",          "B-0001","센터장명","","","","",""),
        ("상황",  "상황실",                          "소방파트/BMS","B-2001","상황실①","","","","",""),
        ("대응반","통제자 (기계파트장)",              "기계파트",   "B-4001","기계파트장","","","","",""),
        ("대응반","대응조 (기계·소방)",               "기계/소방",  "B-4002","기계①","B-1001","소방①","","",""),
        ("지원반","대피조 (건축·보안)",               "건축/보안",  "B-7001","건축①","B-6001","보안①","","",""),
        ("지원반","구조조 (보안)",                    "보안파트",   "B-6002","보안②","","","","",""),
    ],
    "승강기갇힘": [
        ("지휘",  "상황실 (소방·BMS)",               "소방파트/BMS","B-2001","소방①","B-2002","BMS①","","",""),
        ("대응반","전기파트장/승강기안전관리자",        "전기파트",   "B-3001","전기파트장","","","","",""),
        ("대응반","대응조 (OTIS상주·전기파트)",         "전기파트",   "B-3002","전기①","","","","",""),
    ],
    "테러·침입": [
        ("지휘",  "총괄관리자 (센터장)",              "-",          "B-0001","센터장명","","","","",""),
        ("상황",  "상황실 (조장)",                    "소방파트/BMS","B-2001","조장①","","","","",""),
        ("대응반","통제자 (운영파트장)",               "운영파트",   "B-5001","운영파트장","","","","",""),
        ("대응반","보안2조 (서관보안)",                "보안파트",   "B-6001","보안①","B-6002","보안②","","",""),
        ("지원반","응급·지원반 (운영·보안1)",           "운영/보안",  "B-5002","운영①","B-6003","보안③","","",""),
        ("지원반","유도조 (소방파트)",                 "소방파트",   "B-1001","소방①","B-1002","소방②","","",""),
    ],
}

for i, (name, rows) in enumerate(ORG_DATA.items(), 1):
    make_org_sheet(wb, name, i, rows)


# ════════════════════════════════════════════════════════════
# 시트 12 — 상황실 RNR (야간 공통)
# ════════════════════════════════════════════════════════════
ws_rnr = wb.create_sheet("④상황실RNR(야간공통)")
ws_rnr.row_dimensions[1].height = 30
ws_rnr.row_dimensions[2].height = 18
ws_rnr.row_dimensions[3].height = 22

ws_rnr.merge_cells("A1:J1")
c = ws_rnr["A1"]
c.value = "상황실 야간대응 RNR (전 재난 공통)"
c.font = Font(name="맑은 고딕", bold=True, size=14, color="FFFFFF")
c.fill = PatternFill("solid", fgColor="2E4057")
c.alignment = Alignment(horizontal="center", vertical="center")

ws_rnr.merge_cells("A2:J2")
note(ws_rnr, 2, "※ 야간·주말 근무 중 상황 발생 시 상황실 근무자의 역할 분담표")

hdr_cols_rnr = ["반\n(group)", "역할\n(role)", "담당파트", "사번①", "이름①",
                "사번②", "이름②", "사번③", "이름③", "비고"]
for i, h in enumerate(hdr_cols_rnr, 1):
    hdr(ws_rnr.cell(row=3, column=i), h, bg="2E4057", wrap=True)

rnr_rows = [
    ("야간대응","통제자",                    "소방파트/BMS", "B-2001","조장①","","","","","당직조장"),
    ("야간대응","연락·제어",                 "BMS/소방",    "B-2002","BMS①","","","","",""),
    ("야간대응","현장대응 (BMS·전기·보안·운전)","각 파트",    "B-2003","BMS②","B-3001","전기①","B-6001","보안①",""),
    ("야간대응","대피지원 (보안/미화)",        "보안/미화",   "B-6002","보안②","B-9001","미화①","","",""),
]
for r_idx, row_data in enumerate(rnr_rows, 4):
    group = row_data[0]
    bg = GROUP_COLORS.get(group, "E8F0FE")
    for c_idx, val in enumerate(row_data, 1):
        cell(ws_rnr, r_idx, c_idx, val, bg=bg, bold=(c_idx == 1),
             align="center" if c_idx in [1,4,5,6,7,8,9] else "left")

for r_idx in range(4+len(rnr_rows), 4+len(rnr_rows)+3):
    for c_idx in range(1, 11):
        cell(ws_rnr, r_idx, c_idx, "", bg="FAFAFA")

set_col_widths(ws_rnr, [12, 28, 20, 12, 12, 12, 12, 12, 12, 14])
freeze(ws_rnr, "A4")


# ════════════════════════════════════════════════════════════
# 시트 13 — 알람 라우팅 요약 (읽기 전용 참조)
# ════════════════════════════════════════════════════════════
ws_ref = wb.create_sheet("⑤알람라우팅(참조)")
ws_ref.row_dimensions[1].height = 30
ws_ref.row_dimensions[2].height = 18

ws_ref.merge_cells("A1:F1")
c = ws_ref["A1"]
c.value = "알람 라우팅 규칙 요약 (읽기 전용 참조)"
c.font = Font(name="맑은 고딕", bold=True, size=14, color="FFFFFF")
c.fill = PatternFill("solid", fgColor="404040")
c.alignment = Alignment(horizontal="center", vertical="center")

headers_ref = ["시간대", "대상 근무자", "알람 발송 범위", "발송 제외", "근거", "비고"]
for i, h in enumerate(headers_ref, 1):
    hdr(ws_ref.cell(row=2, column=i), h, bg="404040")

ref_rows = [
    ("평일 08:30~17:30","일근자 + 상황실 주간조","해당 재난 조직도 전체","비번자","근무표 shifts + 명부 workType=일근",""),
    ("평일 17:30~익일 08:30","상황실 야간 근무자만","야간 온듀티 조원","비번자·일근자","근무표 shifts 야간 코드",""),
    ("주말·공휴일 전시간","상황실 근무자만","온듀티 조원","비번자·일근자","근무표 shifts","공휴일 별도 플래그 필요"),
    ("긴급 (전체 발령)","전 근무자","명부 active=true 전원","퇴사자","incidents.targetScope=all",""),
    ("감지기 동작 (확인조)","상황실 확인조만","조직도 확인조 역할","일반 대응반","incidents.targetScope=confirm",""),
    ("훈련","지정 조만","incidents.targetTeams 지정","훈련 외 인원","incidents.mode=훈련",""),
]
row_bgs = ["FFFFFF","F5F5F5","FFFFFF","FFF0CC","DDEEFF","E8FFE8"]
for r_idx, (row_data, bg) in enumerate(zip(ref_rows, row_bgs), 3):
    for c_idx, val in enumerate(row_data, 1):
        cell(ws_ref, r_idx, c_idx, val, bg=bg, wrap=True,
             align="center" if c_idx in [1,4] else "left")
    ws_ref.row_dimensions[r_idx].height = 36

set_col_widths(ws_ref, [24, 24, 28, 16, 28, 16])


# ════════════════════════════════════════════════════════════
# 저장
# ════════════════════════════════════════════════════════════
out_path = r"c:\kcoding\twintower-ops\twintower_template.xlsx"
wb.save(out_path)
print(f"[OK] saved: {out_path}")
print(f"     sheets: {len(wb.sheetnames)}")
for s in wb.sheetnames:
    print(f"       - {s}")
