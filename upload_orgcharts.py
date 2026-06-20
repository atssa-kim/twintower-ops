"""
upload_orgcharts.py
orgcharts_skeleton.json (또는 엑셀) → Firestore orgCharts 업로드

사용법:
  python upload_orgcharts.py              # orgcharts_skeleton.json 기준 업로드
  python upload_orgcharts.py --from-excel # 엑셀 → JSON 재생성 후 업로드
  python upload_orgcharts.py --dry-run    # 내용 확인만 (업로드 안 함)
  python upload_orgcharts.py --delete     # 기존 orgCharts 전체 삭제 후 재업로드

Firestore 구조:
  orgCharts/{disasterType}
    .type / .label / .color / .updatedAt
    .groups[]:
      .group
      .roles[]:
        .roleKey      ← taskTemplates 와 동일한 키
        .roleLabel
        .badge
        .depts
        .assignees[]: { empNo, name }
"""

import sys, json, time
from pathlib import Path
from datetime import datetime, timezone
from collections import defaultdict

SA_KEY      = Path(__file__).parent / "serviceAccount.json"
JSON_FILE   = Path(__file__).parent / "orgcharts_skeleton.json"
EXCEL_FILE  = Path(__file__).parent / "orgcharts_input.xlsx"
PROJECT_ID  = "disaster-response-system-f669b"
COLLECTION  = "orgCharts"

DRY_RUN    = "--dry-run"    in sys.argv
FROM_EXCEL = "--from-excel" in sys.argv
DELETE     = "--delete"     in sys.argv


# ── 엑셀 → orgcharts JSON 재생성 ────────────────────────────────────────
def parse_excel() -> dict:
    import openpyxl
    if not EXCEL_FILE.exists():
        print(f"오류: {EXCEL_FILE} 가 없습니다. gen_orgcharts.py 를 먼저 실행하세요.")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_FILE)
    now_ms = int(datetime.now(timezone.utc).timestamp() * 1000)

    # 기존 skeleton 로드하여 meta 정보 유지
    base = json.loads(JSON_FILE.read_text(encoding="utf-8")) if JSON_FILE.exists() else {}
    base_disasters = base.get("disasters", {})

    disasters = {}
    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("조직도_"):
            continue
        d_key = sheet_name.replace("조직도_", "")
        ws = wb[sheet_name]

        base_d = base_disasters.get(d_key, {})
        groups_dict = defaultdict(lambda: defaultdict(
            lambda: {"roleLabel": "", "badge": "", "depts": "", "assignees": []}
        ))

        # 4행부터 데이터 읽기
        for row in ws.iter_rows(min_row=4, values_only=True):
            group, role_key, role_label, badge, depts, emp_no, name, *_ = \
                (list(row) + [""] * 8)[:8]

            if not role_key or role_key == "roleKey":
                continue
            if not emp_no or emp_no in ("B-XXXX", ""):
                continue   # 미입력 행 건너뜀

            if not groups_dict[group][role_key]["roleLabel"]:
                groups_dict[group][role_key]["roleLabel"] = role_label or ""
                groups_dict[group][role_key]["badge"]     = badge     or ""
                groups_dict[group][role_key]["depts"]     = depts     or ""

            groups_dict[group][role_key]["assignees"].append(
                {"empNo": str(emp_no).strip(), "name": str(name).strip()}
            )

        groups = []
        for grp, roles_dict in groups_dict.items():
            roles = []
            for rk, rv in roles_dict.items():
                roles.append({
                    "roleKey":   rk,
                    "roleLabel": rv["roleLabel"],
                    "badge":     rv["badge"],
                    "depts":     rv["depts"],
                    "assignees": rv["assignees"],
                })
            groups.append({"group": grp, "roles": roles})

        disasters[d_key] = {
            "type":      d_key,
            "label":     base_d.get("label", d_key),
            "color":     base_d.get("color", ""),
            "updatedAt": now_ms,
            "groups":    groups,
        }

    return {"disasters": disasters}


# ── 미리보기 ─────────────────────────────────────────────────────────────
def preview(data: dict):
    print(f"\n{'재난':10s}  {'반':12s}  {'roleKey':25s}  담당자수")
    print("-" * 68)
    total_assignees = 0
    for d_key, disaster in data["disasters"].items():
        for grp_obj in disaster.get("groups", []):
            for role_obj in grp_obj.get("roles", []):
                n = len(role_obj.get("assignees", []))
                total_assignees += n
                d_safe  = d_key.encode("cp949","replace").decode("cp949")
                rk_safe = role_obj["roleKey"].encode("cp949","replace").decode("cp949")
                g_safe  = grp_obj["group"].encode("cp949","replace").decode("cp949")
                print(f"  {d_safe:8s}  {g_safe:10s}  {rk_safe:23s}  {n}명")
    print("-" * 68)
    print(f"  총 담당자 배정: {total_assignees}건\n")


# ── 서비스 계정 키 확인 ──────────────────────────────────────────────────
def check_sa_key():
    if SA_KEY.exists():
        return True
    print()
    print("=" * 60)
    print("  serviceAccount.json 파일이 없습니다.")
    print(f"  Firebase 콘솔 → 프로젝트 설정 → 서비스 계정")
    print(f"  → '새 비공개 키 생성' → serviceAccount.json 으로 저장")
    print("=" * 60)
    return False


# ── Firestore 삭제 ───────────────────────────────────────────────────────
def delete_collection(col_ref, batch_size=300):
    docs = list(col_ref.limit(batch_size).stream())
    deleted = 0
    for doc in docs:
        doc.reference.delete()
        deleted += 1
    return deleted


# ── 업로드 ───────────────────────────────────────────────────────────────
def upload(db, data: dict):
    now_ms = int(datetime.now(timezone.utc).timestamp() * 1000)
    col = db.collection(COLLECTION)

    for d_key, disaster in data["disasters"].items():
        doc_data = {
            "type":      disaster.get("type", d_key),
            "label":     disaster.get("label", d_key),
            "color":     disaster.get("color", ""),
            "updatedAt": disaster.get("updatedAt", now_ms),
            "groups":    disaster.get("groups", []),
        }
        col.document(d_key).set(doc_data)
        n_roles = sum(len(g["roles"]) for g in doc_data["groups"])
        n_assign = sum(
            len(r["assignees"])
            for g in doc_data["groups"]
            for r in g["roles"]
        )
        d_safe = d_key.encode("cp949","replace").decode("cp949")
        print(f"  [{d_safe}] roles={n_roles}, assignees={n_assign} 업로드 완료")
        time.sleep(0.05)

    print("\n  orgCharts 업로드 완료.")
    print(f"  Firestore 확인:")
    print(f"  https://console.firebase.google.com/project/{PROJECT_ID}"
          f"/firestore/data/{COLLECTION}")


# ── 메인 ────────────────────────────────────────────────────────────────
def main():
    # 1. 데이터 로드
    if FROM_EXCEL:
        print("엑셀에서 데이터 읽는 중...")
        data = parse_excel()
        # 엑셀 내용을 JSON으로도 저장
        JSON_FILE.write_text(
            json.dumps(data, ensure_ascii=False, indent=2),
            encoding="utf-8"
        )
        print(f"  → {JSON_FILE} 재생성 완료")
    else:
        if not JSON_FILE.exists():
            print(f"오류: {JSON_FILE} 없음. gen_orgcharts.py 를 먼저 실행하세요.")
            sys.exit(1)
        data = json.loads(JSON_FILE.read_text(encoding="utf-8"))
        if "disasters" not in data:
            print("오류: orgcharts_skeleton.json 형식이 잘못됐습니다.")
            sys.exit(1)

    # 2. 미리보기
    preview(data)

    if DRY_RUN:
        print("[DRY-RUN] 실제 업로드는 하지 않습니다.")
        return

    # 3. 서비스 계정 키 확인
    if not check_sa_key():
        sys.exit(1)

    # 4. Firebase 초기화
    import firebase_admin
    from firebase_admin import credentials, firestore

    if not firebase_admin._apps:
        cred = credentials.Certificate(str(SA_KEY))
        firebase_admin.initialize_app(cred, {"projectId": PROJECT_ID})
    db = firestore.client()

    # 5. 삭제 옵션
    if DELETE:
        print("기존 orgCharts 삭제 중...")
        n = delete_collection(db.collection(COLLECTION))
        print(f"  {n}개 문서 삭제\n")

    # 6. 업로드
    upload(db, data)


if __name__ == "__main__":
    main()
