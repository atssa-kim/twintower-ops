"""
gen_orgcharts.py
tasks.json 의 roleKey 를 기반으로 orgcharts_skeleton.json 을 생성한다.
사용자는 생성된 JSON 의 assignees[] 에 실제 사번·이름을 입력한 뒤
upload_orgcharts.py 로 Firestore 에 올린다.

출력 구조 (orgcharts_skeleton.json):
{
  "disasters": {
    "화재": {
      "type": "화재",
      "label": "...",
      "groups": [
        {
          "group": "지휘",
          "roles": [
            {
              "roleKey":   "지휘/총괄",
              "roleLabel": "🎖️ 총괄자 (센터장)",
              "badge":     "총괄",
              "depts":     "소방파트 등 담당 파트 메모",
              "assignees": [
                { "empNo": "B-XXXX", "name": "이름 입력" }
              ]
            }
          ]
        }
      ]
    }
  }
}
"""

import json, re
from pathlib import Path
from datetime import datetime, timezone
from collections import defaultdict

TASKS_JSON = Path(__file__).parent / "tasks.json"
OUT_JSON   = Path(__file__).parent / "orgcharts_skeleton.json"
OUT_EXCEL  = Path(__file__).parent / "orgcharts_input.xlsx"

# 재난별 역할-담당파트 힌트 (엑셀 조직도에서 가져온 참고 정보)
DEPT_HINTS = {
    "화재": {
        "지휘/총괄":    "-",
        "지휘/통제":    "소방파트",
        "연락반/상황":  "소방파트/BMS",
        "대응반/출동":  "전기파트/소방파트",
        "대응반/소화":  "기계파트",
        "대응반/구조":  "보안파트",
        "유도반/유도":  "운영파트/보안파트/주차파트",
        "유도반/경계":  "보안파트",
        "복구반/복구":  "미화파트/시설",
    },
    "정전": {
        "지휘/총괄":    "-",
        "연락/상황":    "소방파트/BMS",
        "대응반/통제":  "전기파트",
        "대응반/대응":  "전기파트",
        "지원반/지원":  "소방파트/기계파트",
        "지원반/유도":  "운영파트/보안파트",
    },
    "홍수": {
        "총괄관리자/총괄": "-",
        "상황실/상황":     "소방파트/BMS",
        "대응반/응급":     "기계파트/소방파트",
        "대응반/복구E":    "전기파트",
        "대응반/복구B":    "건축파트/미화파트",
        "대피반/유도":     "보안파트",
        "대피반/보안":     "운영파트",
        "지원반/의료":     "보안파트",
        "지원반/관리":     "운영파트",
    },
    "태풍": {
        "지휘/총괄":   "-",
        "상황/상황":   "소방파트/BMS",
        "대응반/통제": "건축파트",
        "대응반/대응": "건축파트/기계파트/전기파트/운영파트",
        "지원반/지원": "미화파트/보안파트/주차파트",
        "지원반/구조": "보안파트",
    },
    "폭설": {
        "지휘/총괄":    "-",
        "대응반/대응1": "운영파트/소방파트/건축파트",
        "대응반/대응2": "전기파트/기계파트/품질파트",
        "지원반/지원1": "미화파트",
        "지원반/지원2": "보안파트/주차파트",
    },
    "지진": {
        "지휘/총괄":    "-",
        "대응반/대응1": "전기파트/건축파트",
        "대응반/대응2": "기계파트/미화파트/주차파트",
        "대피반/대피":  "보안파트",
        "지원반/지원":  "운영파트",
    },
    "가스누출": {
        "지휘/총괄":   "-",
        "상황/상황":   "소방파트/BMS",
        "대응반/통제": "기계파트",
        "대응반/대응": "기계파트/소방파트",
        "지원반/대피": "건축파트/보안파트",
        "지원반/구조": "보안파트",
    },
    "승강기": {
        "지휘/상황":   "소방파트/BMS",
        "대응반/전기": "전기파트",
        "대응반/대응": "전기파트/OTIS",
    },
    "테러": {
        "지휘/총괄":   "-",
        "상황/상황":   "소방파트/BMS",
        "대응반/통제": "운영파트",
        "대응반/보안2":"보안파트",
        "지원반/응급": "운영파트/보안파트",
        "지원반/유도": "소방파트",
    },
}


def group_roles_by_group(roles: list) -> list:
    """roles 리스트를 group 기준으로 묶어 반환."""
    groups_dict = defaultdict(list)
    for r in roles:
        groups_dict[r["group"]].append(r)
    result = []
    for grp, role_list in groups_dict.items():
        result.append({
            "group": grp,
            "roles": [
                {
                    "roleKey":   r["roleKey"],
                    "roleLabel": r["roleLabel"],
                    "badge":     r["badge"],
                    "depts":     DEPT_HINTS.get(
                                    # key는 tasks.json의 disaster key
                                    "_current_key_",
                                    {}
                                 ).get(r["roleKey"], "파트 입력"),
                    "assignees": [
                        {"empNo": "B-XXXX", "name": "이름 입력"}
                    ]
                }
                for r in role_list
            ]
        })
    return result


def main():
    data = json.loads(TASKS_JSON.read_text(encoding="utf-8"))
    now_ms = int(datetime.now(timezone.utc).timestamp() * 1000)

    out = {
        "version":     data["version"],
        "generatedAt": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "note":        "assignees[] 에 실제 사번(empNo)과 이름(name)을 입력 후 upload_orgcharts.py 실행",
        "disasters": {}
    }

    for key, disaster in data["disasters"].items():
        hints = DEPT_HINTS.get(key, {})

        # 역할 리스트를 그룹별로 묶기
        groups_dict = defaultdict(list)
        for r in disaster["roles"]:
            dept_hint = hints.get(r["roleKey"], "파트 입력")
            groups_dict[r["group"]].append({
                "roleKey":   r["roleKey"],
                "roleLabel": r["roleLabel"],
                "badge":     r["badge"],
                "depts":     dept_hint,
                "assignees": [
                    {"empNo": "B-XXXX", "name": "이름 입력"}
                ]
            })

        groups = [
            {"group": g, "roles": roles}
            for g, roles in groups_dict.items()
        ]

        entry = {
            "type":       key,
            "label":      disaster["label"],
            "color":      disaster.get("color", ""),
            "updatedAt":  now_ms,
            "groups":     groups,
        }

        # 화재 추가 변형 (야간/시나리오 메모)
        if key == "화재":
            sc_keys = [k for k in data["disasters"]["화재"].get("scenarios", {})
                       if k != "general"]
            entry["scenarioKeys"] = sc_keys
            entry["note"] = (
                "시나리오별 조직도는 기본 화재와 동일 인원이므로 "
                "별도 orgCharts 문서는 선택 사항입니다."
            )

        out["disasters"][key] = entry

    OUT_JSON.write_text(
        json.dumps(out, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )
    print(f"[OK] {OUT_JSON}")
    print(f"     {len(out['disasters'])} disasters written")

    # ── Excel 입력표 생성 ────────────────────────────────────────────
    _make_excel(out)
    print(f"[OK] {OUT_EXCEL}")
    print()
    print("  [다음 단계]")
    print("  1. orgcharts_input.xlsx 를 열어 사번/이름 입력")
    print("     (또는 orgcharts_skeleton.json 의 assignees 직접 편집)")
    print("  2. python upload_orgcharts.py --from-excel")
    print("     (엑셀에서 JSON 재생성 + Firestore 업로드)")


# ── Excel 입력표 생성 ──────────────────────────────────────────────────
def _make_excel(out: dict):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    def side(s="thin"):
        return Side(style=s, color="AAAAAA")
    BD = Border(left=side(), right=side(), top=side(), bottom=side())
    BD2 = Border(left=side("medium"), right=side("medium"),
                 top=side("medium"), bottom=side("medium"))

    def hcell(c, v, bg="1F4E79", fg="FFFFFF", sz=10, bold=True):
        c.value = v
        c.font = Font(name="맑은 고딕", bold=bold, color=fg, size=sz)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = BD2

    def dcell(c, v, bg="FFFFFF", bold=False, align="left"):
        c.value = v
        c.font = Font(name="맑은 고딕", bold=bold, size=10)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal=align, vertical="center")
        c.border = BD

    GROUP_BG = {
        "지휘": "FFF2CC", "총괄관리자": "FFF2CC",
        "연락반": "DDEBF7", "연락": "DDEBF7",
        "상황": "DDEBF7", "상황실": "DDEBF7",
        "대응반": "FCE4D6",
        "유도반": "E2EFDA", "대피반": "E2EFDA",
        "복구반": "F2F2F2", "지원반": "F2F2F2",
    }
    DISASTER_BG = {
        "화재": "C00000", "정전": "7B6100", "홍수": "1F4E79",
        "태풍": "1F4E79", "폭설": "2E4057", "지진": "5C3A00",
        "가스누출": "375623", "승강기": "4A235A", "테러": "404040",
    }

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for d_key, disaster in out["disasters"].items():
        ws = wb.create_sheet(f"조직도_{d_key}")
        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 18
        ws.row_dimensions[3].height = 22

        title_bg = DISASTER_BG.get(d_key, "1F4E79")
        ws.merge_cells("A1:H1")
        hcell(ws["A1"],
              f"조직도 — {d_key}  (empNo·name 을 실제 값으로 교체하세요)",
              bg=title_bg, sz=13)

        ws.merge_cells("A2:H2")
        ws["A2"].value = (
            "※ roleKey 는 수정 금지.  사번(empNo)과 이름(name)만 입력."
            "  담당자가 여러 명이면 행을 추가(roleKey·group 복사)."
        )
        ws["A2"].font = Font(name="맑은 고딕", italic=True,
                             color="555555", size=9)
        ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

        headers = ["반(group)", "roleKey", "역할명(roleLabel)",
                   "badge", "담당파트", "사번(empNo)", "이름(name)", "비고"]
        for i, h in enumerate(headers, 1):
            hcell(ws.cell(3, i), h, bg=title_bg)

        row = 4
        for grp_obj in disaster["groups"]:
            grp = grp_obj["group"]
            bg  = GROUP_BG.get(grp, "FFFFFF")
            for role_obj in grp_obj["roles"]:
                for assignee in role_obj["assignees"]:
                    dcell(ws.cell(row, 1), grp,  bg=bg, bold=True,  align="center")
                    dcell(ws.cell(row, 2), role_obj["roleKey"],   bg=bg, align="center")
                    dcell(ws.cell(row, 3), role_obj["roleLabel"], bg=bg)
                    dcell(ws.cell(row, 4), role_obj["badge"],     bg=bg, align="center")
                    dcell(ws.cell(row, 5), role_obj["depts"],     bg=bg)
                    dcell(ws.cell(row, 6), assignee["empNo"],     bg=bg, align="center")
                    dcell(ws.cell(row, 7), assignee["name"],      bg=bg)
                    dcell(ws.cell(row, 8), "",                    bg=bg)
                    row += 1

        # 빈 입력 행
        for _ in range(3):
            for c in range(1, 9):
                dcell(ws.cell(row, c), "", bg="FAFAFA")
            row += 1

        widths = [12, 22, 28, 8, 22, 14, 12, 12]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(i)
            ].width = w
        ws.freeze_panes = "A4"

    wb.save(OUT_EXCEL)


if __name__ == "__main__":
    main()
