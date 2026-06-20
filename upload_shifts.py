"""
upload_shifts.py
twintower_template.xlsx 의 '②근무표(shifts)' 시트를 읽어
Firestore shifts/{YYYY-MM}/members/{empNo} 로 업로드한다.

Firestore 구조:
  shifts/{YYYY-MM}/members/{empNo}
    .empNo       : "B-1001"
    .name        : "김상백"
    .dept        : "소방파트"
    .shiftGroup  : "소방1조"
    .schedule    : { "2026-06-16": "주", "2026-06-17": "야", ... }
    .updatedAt   : timestamp(ms)

사용법:
  python upload_shifts.py                      # 엑셀에서 연·월 자동 감지 후 업로드
  python upload_shifts.py --year 2026 --month 7   # 특정 연·월 지정
  python upload_shifts.py --dry-run            # 내용 확인만
  python upload_shifts.py --delete             # 해당 월 삭제 후 재업로드
"""

import sys, json, re
from pathlib import Path
from datetime import datetime, timezone, date
import calendar

SA_KEY      = Path(__file__).parent / "serviceAccount.json"
EXCEL_FILE  = Path(__file__).parent / "twintower_template.xlsx"
PROJECT_ID  = "disaster-response-system-f669b"
COLLECTION  = "shifts"
SHEET_NAME  = "②근무표(shifts)"

DRY_RUN = "--dry-run" in sys.argv
DELETE  = "--delete"  in sys.argv

# CLI 연·월 파싱
def get_year_month():
    y = m = None
    for i, a in enumerate(sys.argv):
        if a == "--year"  and i+1 < len(sys.argv): y = int(sys.argv[i+1])
        if a == "--month" and i+1 < len(sys.argv): m = int(sys.argv[i+1])
    return y, m


# ── 엑셀 파싱 ────────────────────────────────────────────────────────────
def parse_excel(year_hint=None, month_hint=None):
    import openpyxl
    if not EXCEL_FILE.exists():
        print(f"오류: {EXCEL_FILE} 없음. make_template.py 를 먼저 실행하세요.")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    # 시트 이름 유연 탐색 (번호·이름 부분 일치)
    ws = None
    for name in wb.sheetnames:
        if "근무표" in name or "shifts" in name.lower():
            ws = wb[name]
            break
    if ws is None:
        print(f"오류: '{SHEET_NAME}' 시트를 찾을 수 없습니다.")
        print(f"  보유 시트: {wb.sheetnames}")
        sys.exit(1)

    # ── 연·월 감지 (제목 행 A1 에서 "2026년 __월" 파싱) ──────────────
    title = str(ws["A1"].value or "")
    year, month = year_hint, month_hint
    if not (year and month):
        m = re.search(r'(\d{4})[년\s]+(\d{1,2})[월\s]', title)
        if m:
            year  = year  or int(m.group(1))
            month = month or int(m.group(2))
        else:
            now   = datetime.now()
            year  = year  or now.year
            month = month or now.month
            print(f"  [INFO] 연·월 자동 감지 실패 → 현재 연월 사용: {year}-{month:02d}")

    ym_str   = f"{year}-{month:02d}"
    days_in_month = calendar.monthrange(year, month)[1]

    # ── 날짜 헤더 행(3행) 에서 컬럼→날짜 매핑 ──────────────────────────
    # 고정 컬럼: A=사번, B=이름, C=파트, D=교대조 (1~4)
    # 5번 컬럼부터 날짜
    col_to_date = {}
    for col in range(5, 5 + days_in_month + 5):
        cell_val = ws.cell(row=3, column=col).value
        if cell_val is None:
            continue
        try:
            day = int(cell_val)
            if 1 <= day <= days_in_month:
                d = date(year, month, day)
                col_to_date[col] = d.strftime("%Y-%m-%d")
        except (ValueError, TypeError):
            pass

    if not col_to_date:
        print("오류: 날짜 헤더(3행)를 찾을 수 없습니다.")
        sys.exit(1)

    # ── 데이터 행(5행~) 파싱 ────────────────────────────────────────────
    VALID_CODES = {"주", "야", "아", "비"}
    members = []

    for row in ws.iter_rows(min_row=5, values_only=True):
        emp_no, name, dept, grp = (list(row) + ["","","",""])[:4]
        if not emp_no or str(emp_no).strip() in ("사번", "B-XXXX", ""):
            continue
        emp_no = str(emp_no).strip()
        # 유효한 사번 패턴만 허용 (예: B-1001, 1001, EMP-001 등)
        if not re.match(r'^[A-Za-z0-9][A-Za-z0-9\-_]{0,19}$', emp_no):
            continue
        name   = str(name   or "").strip()
        dept   = str(dept   or "").strip()
        grp    = str(grp    or "").strip()

        schedule = {}
        for col, date_str in col_to_date.items():
            code = row[col - 1] if col - 1 < len(row) else None
            code = str(code).strip() if code else "비"
            if code not in VALID_CODES:
                code = "비"
            schedule[date_str] = code

        if schedule:
            members.append({
                "empNo":      emp_no,
                "name":       name,
                "dept":       dept,
                "shiftGroup": grp,
                "schedule":   schedule,
            })

    return ym_str, members


# ── 미리보기 ─────────────────────────────────────────────────────────────
def preview(ym_str, members):
    days = list(members[0]["schedule"].keys()) if members else []
    print(f"\n  대상 연월: {ym_str}  ({len(days)}일)")
    print(f"  업로드 인원: {len(members)}명\n")
    print(f"  {'사번':12s} {'이름':8s} {'교대조':10s} 근무코드 샘플(처음 7일)")
    print("  " + "-" * 60)
    for m in members:
        sample = " ".join(list(m["schedule"].values())[:7])
        emp_safe  = m['empNo'].encode('cp949','replace').decode('cp949')
        name_safe = m['name'].encode('cp949','replace').decode('cp949')
        grp_safe  = m['shiftGroup'].encode('cp949','replace').decode('cp949')
        print(f"  {emp_safe:12s} {name_safe:8s} {grp_safe:10s} {sample}")
    print()


# ── 서비스 계정 키 확인 ──────────────────────────────────────────────────
def check_sa_key():
    if SA_KEY.exists():
        return True
    print()
    print("=" * 60)
    print("  serviceAccount.json 파일이 없습니다.")
    print(f"  Firebase 콘솔 -> 프로젝트 설정 -> 서비스 계정")
    print(f"  -> '새 비공개 키 생성' -> serviceAccount.json 으로 저장")
    print("=" * 60)
    return False


# ── Firestore 월 데이터 삭제 ─────────────────────────────────────────────
def delete_month(db, ym_str):
    col = db.collection(COLLECTION).document(ym_str).collection("members")
    docs = list(col.stream())
    for d in docs:
        d.reference.delete()
    return len(docs)


# ── 업로드 ───────────────────────────────────────────────────────────────
def upload(db, ym_str, members):
    from google.cloud import firestore as _fs

    now_ms  = int(datetime.now(timezone.utc).timestamp() * 1000)
    col     = db.collection(COLLECTION).document(ym_str).collection("members")
    batch   = db.batch()
    count   = 0

    # 월 메타 문서
    db.collection(COLLECTION).document(ym_str).set({
        "yearMonth": ym_str,
        "memberCount": len(members),
        "updatedAt": now_ms,
    })

    for m in members:
        ref = col.document(m["empNo"])
        batch.set(ref, {
            "empNo":      m["empNo"],
            "name":       m["name"],
            "dept":       m["dept"],
            "shiftGroup": m["shiftGroup"],
            "schedule":   m["schedule"],
            "updatedAt":  now_ms,
        })
        count += 1
        if count % 400 == 0:
            batch.commit()
            batch = db.batch()

    if count % 400 != 0:
        batch.commit()

    print(f"  [{ym_str}] {count}명 업로드 완료")
    print(f"  Firestore 확인:")
    print(f"  https://console.firebase.google.com/project/{PROJECT_ID}"
          f"/firestore/data/{COLLECTION}/{ym_str}")


# ── 메인 ────────────────────────────────────────────────────────────────
def main():
    year_hint, month_hint = get_year_month()

    # 1. 엑셀 파싱
    ym_str, members = parse_excel(year_hint, month_hint)

    if not members:
        print("입력된 교대 근무자가 없습니다. 엑셀을 확인하세요.")
        sys.exit(1)

    # 2. 미리보기
    preview(ym_str, members)

    if DRY_RUN:
        print("[DRY-RUN] 실제 업로드는 하지 않습니다.")
        return

    # 3. 키 확인
    if not check_sa_key():
        sys.exit(1)

    # 4. Firebase 초기화
    import firebase_admin
    from firebase_admin import credentials, firestore

    if not firebase_admin._apps:
        cred = credentials.Certificate(str(SA_KEY))
        firebase_admin.initialize_app(cred, {"projectId": PROJECT_ID})
    db = firestore.client()

    # 5. 삭제 옵션
    if DELETE:
        n = delete_month(db, ym_str)
        print(f"  기존 {ym_str} 데이터 {n}건 삭제\n")

    # 6. 업로드
    upload(db, ym_str, members)

    # 7. 다음 달 안내
    y, m = int(ym_str[:4]), int(ym_str[5:])
    next_m = m + 1 if m < 12 else 1
    next_y = y if m < 12 else y + 1
    print(f"\n  다음 달 업로드:")
    print(f"  python upload_shifts.py --year {next_y} --month {next_m}")


if __name__ == "__main__":
    main()
